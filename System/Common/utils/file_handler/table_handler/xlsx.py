# -*- encoding: utf-8 -*-
"""
@File Name      :   xlsx.py    
@Create Time    :   2021/12/1 14:58
@Description    :   
@Version        :   
@License        :   
@Author         :   diklios
@Contact Email  :   diklios5768@gmail.com
@Github         :   https://github.com/diklios5768
@Blog           :   
@Motto          :   All our science, measured against reality, is primitive and childlike - and yet it is the most precious thing we have.
"""
__auth__ = 'diklios'

import os
from io import BytesIO

import pandas as pd
from django.conf import settings
from openpyxl import Workbook

from Common.utils.file_handler.dir import make_dir


def generate_xlsx_file(filename, table_sheets, file_dir):
    wb = Workbook()
    for table_sheet in table_sheets:
        ws = wb.create_sheet(table_sheet['sheet_name'])
        table_data = table_sheet['sheet_data']
        rows_length = len(table_data)
        for i in range(rows_length):
            col_length = len(table_data[i])
            for j in range(col_length):
                ws.cell(row=i + 1, column=j + 1, value=table_data[i][j])
    wb.remove(wb['Sheet'])
    if file_dir is not None:
        if make_dir(file_dir):
            if not file_dir.endswith('/'):
                file_dir += '/'
            wb.save(file_dir + filename)
        else:
            print('make dir fail')
    else:
        wb.save(os.path.join(settings.BASE_DIR, filename))
    return True


def generate_xlsx_file_io(df: pd.DataFrame):
    file = BytesIO()
    # By setting the 'engine' in the ExcelWriter constructor.
    writer = pd.ExcelWriter(file, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    # Save the workbook
    writer.save()
    # Seek to the beginning and read to copy the workbook to a variable in memory
    file.seek(0)
    return file
